"""REST views for contract attachments and export."""
import io
from decimal import Decimal

from django.http import HttpResponse, JsonResponse, FileResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from apps.core.permissions import get_current_user_from_request
from .models import Contract, ContractAttachment


@method_decorator(csrf_exempt, name="dispatch")
class AttachmentDownloadView(View):
    """REST endpoint for downloading contract attachments."""

    def get(self, request, attachment_id):
        """
        Download a contract attachment.

        Requires authentication and verifies tenant ownership.
        """
        # Authenticate
        user = get_current_user_from_request(request)
        if not user:
            return JsonResponse({"error": "Authentication required"}, status=401)

        if not user.tenant:
            return JsonResponse({"error": "No tenant assigned"}, status=403)

        # Find attachment (with tenant verification)
        attachment = ContractAttachment.objects.filter(
            tenant=user.tenant,
            id=attachment_id,
        ).first()

        if not attachment:
            return JsonResponse({"error": "Attachment not found"}, status=404)

        # Serve file
        try:
            response = FileResponse(
                attachment.file.open("rb"),
                content_type=attachment.content_type,
            )
            # Check if preview mode (inline viewing) is requested
            preview = request.GET.get("preview", "").lower() in ("true", "1")
            disposition = "inline" if preview else "attachment"
            response["Content-Disposition"] = f'{disposition}; filename="{attachment.original_filename}"'
            response["Content-Length"] = attachment.file_size
            return response
        except FileNotFoundError:
            return JsonResponse({"error": "File not found on storage"}, status=404)


@method_decorator(csrf_exempt, name="dispatch")
class ContractExportView(View):
    """REST endpoint for exporting active contracts as Excel."""

    def get(self, request):
        user = get_current_user_from_request(request)
        if not user:
            return JsonResponse({"error": "Authentication required"}, status=401)

        if not user.has_perm_check("contracts", "export"):
            return JsonResponse({"error": "Permission denied"}, status=403)

        language = request.GET.get("language", "de")
        if language not in ("de", "en"):
            language = "de"

        contracts = (
            Contract.objects.filter(
                tenant=user.tenant,
                status=Contract.Status.ACTIVE,
            )
            .select_related("customer")
            .prefetch_related("items__product", "items__price_periods")
            .order_by("customer__name", "name")
        )

        if not contracts.exists():
            return JsonResponse(
                {"error": "No active contracts found"}, status=404
            )

        content = self._generate_excel(
            list(contracts), user.tenant, language
        )
        filename = "contracts-export.xlsx"
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response = HttpResponse(content, content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(content)
        return response

    def _generate_excel(self, contracts, tenant, language):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        h_de = {
            "customer": "Kunde",
            "sales_order": "SO-Nummer",
            "contract_number": "Vertrag",
            "po_number": "Bestellnummer",
            "ab_number": "AB-Nummer",
            "invoicing_instructions": "Rechnungshinweise",
            "contract_start": "Vertragsbeginn",
            "contract_end": "Vertragsende",
            "billing_schedule": "Abrechnungsintervall",
            "monthly_rate": "Monatliche Rate",
            "arr": "ARR",
            "item": "Position",
            "item_description": "Beschreibung",
            "item_start": "Position gültig ab",
            "item_billing_start": "Abrechnung ab",
            "item_billing_end": "Abrechnung bis",
            "quantity": "Menge",
            "unit_price": "Einzelpreis",
            "price_period": "Preisperiode",
            "amount": "Betrag",
            "is_one_off": "Einmalig",
            "total": "Gesamtbetrag",
            "summary_title": "Vertragsübersicht",
        }
        h_en = {
            "customer": "Customer",
            "sales_order": "Sales Order",
            "contract_number": "Contract",
            "po_number": "PO Number",
            "ab_number": "Order Confirmation",
            "invoicing_instructions": "Invoicing Instructions",
            "contract_start": "Contract Start",
            "contract_end": "Contract End",
            "billing_schedule": "Billing Schedule",
            "monthly_rate": "Monthly Rate",
            "arr": "ARR",
            "item": "Item",
            "item_description": "Description",
            "item_start": "Item Effective Date",
            "item_billing_start": "Billing Start",
            "item_billing_end": "Billing End",
            "quantity": "Quantity",
            "unit_price": "Unit Price",
            "price_period": "Price Period",
            "amount": "Amount",
            "is_one_off": "One-off",
            "total": "Total",
            "summary_title": "Contract Overview",
        }
        h = h_de if language == "de" else h_en

        interval_names = {
            "monthly": "Monthly" if language == "en" else "Monatlich",
            "quarterly": "Quarterly" if language == "en" else "Vierteljährlich",
            "semi_annual": "Semi-annual" if language == "en" else "Halbjährlich",
            "annual": "Annual" if language == "en" else "Jährlich",
            "biennial": "2 Years" if language == "en" else "2 Jahre",
            "triennial": "3 Years" if language == "en" else "3 Jahre",
            "quadrennial": "4 Years" if language == "en" else "4 Jahre",
            "quinquennial": "5 Years" if language == "en" else "5 Jahre",
        }

        price_period_names = {
            "monthly": "Monthly" if language == "en" else "Monatlich",
            "bi_monthly": "2 Months" if language == "en" else "2 Monate",
            "quarterly": "Quarterly" if language == "en" else "Vierteljährlich",
            "semi_annual": "Semi-annual" if language == "en" else "Halbjährlich",
            "annual": "Annual" if language == "en" else "Jährlich",
            "biennial": "2 Years" if language == "en" else "2 Jahre",
            "triennial": "3 Years" if language == "en" else "3 Jahre",
            "quadrennial": "4 Years" if language == "en" else "4 Jahre",
            "quinquennial": "5 Years" if language == "en" else "5 Jahre",
        }

        wb = Workbook()
        currency = tenant.currency
        currency_format = f'#,##0.00 "{currency}"'

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2563EB", end_color="2563EB", fill_type="solid"
        )
        total_fill = PatternFill(
            start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"
        )
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ---- Summary Sheet ----
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = h["summary_title"]
        ws_summary["A1"].font = title_font

        summary_headers = [
            h["customer"],
            h["sales_order"],
            h["contract_number"],
            h["po_number"],
            h["invoicing_instructions"],
            h["contract_start"],
            h["contract_end"],
            h["billing_schedule"],
            h["monthly_rate"],
            h["arr"],
        ]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row = 4
        total_monthly = Decimal("0")
        total_arr = Decimal("0")

        for contract in contracts:
            customer_name = contract.customer.name
            if contract.customer.netsuite_customer_number:
                customer_name = (
                    f"{contract.customer.netsuite_customer_number} {customer_name}"
                )

            # Calculate monthly rate from recurring items
            monthly = Decimal("0")
            for item in contract.items.all():
                if item.is_one_off:
                    continue
                monthly += item.monthly_unit_price * item.quantity

            arr = monthly * 12

            ws_summary.cell(row=row, column=1, value=customer_name)
            ws_summary.cell(
                row=row, column=2, value=contract.netsuite_sales_order_number or ""
            )
            ws_summary.cell(
                row=row, column=3, value=contract.netsuite_contract_number or ""
            )
            ws_summary.cell(row=row, column=4, value=contract.po_number or "")
            ws_summary.cell(row=row, column=5, value=contract.invoice_text or "")
            ws_summary.cell(row=row, column=6, value=contract.start_date)
            ws_summary.cell(row=row, column=7, value=contract.end_date)
            ws_summary.cell(
                row=row,
                column=8,
                value=interval_names.get(
                    contract.billing_interval, contract.billing_interval
                ),
            )

            mc = ws_summary.cell(row=row, column=9, value=float(monthly))
            mc.number_format = currency_format

            ac = ws_summary.cell(row=row, column=10, value=float(arr))
            ac.number_format = currency_format

            total_monthly += monthly
            total_arr += arr
            row += 1

        # Total row
        ws_summary.cell(row=row, column=1, value=h["total"]).font = Font(bold=True)
        for col in range(1, 9):
            ws_summary.cell(row=row, column=col).fill = total_fill

        mc = ws_summary.cell(row=row, column=9, value=float(total_monthly))
        mc.number_format = currency_format
        mc.font = Font(bold=True)
        mc.fill = total_fill

        ac = ws_summary.cell(row=row, column=10, value=float(total_arr))
        ac.number_format = currency_format
        ac.font = Font(bold=True)
        ac.fill = total_fill

        summary_widths = [35, 18, 30, 15, 40, 14, 14, 18, 15, 15]
        for col, width in enumerate(summary_widths, 1):
            ws_summary.column_dimensions[get_column_letter(col)].width = width

        # ---- Details Sheet ----
        ws_details = wb.create_sheet("Details")
        detail_headers = [
            h["customer"],
            h["sales_order"],
            h["contract_number"],
            h["po_number"],
            h["ab_number"],
            h["item"],
            h["item_description"],
            h["invoicing_instructions"],
            h["contract_start"],
            h["contract_end"],
            h["item_start"],
            h["item_billing_start"],
            h["item_billing_end"],
            h["billing_schedule"],
            h["quantity"],
            h["unit_price"],
            h["price_period"],
            h["amount"],
            h["is_one_off"],
        ]
        for col, header in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row = 2
        for contract in contracts:
            customer_name = contract.customer.name
            if contract.customer.netsuite_customer_number:
                customer_name = (
                    f"{contract.customer.netsuite_customer_number} {customer_name}"
                )
            billing_schedule = interval_names.get(
                contract.billing_interval, contract.billing_interval
            )

            for item in contract.items.all():
                product_name = item.product.name if item.product else ""
                ab_number = (
                    item.order_confirmation_number
                    or contract.order_confirmation_number
                    or ""
                )
                total_amount = item.unit_price * item.quantity

                ws_details.cell(row=row, column=1, value=customer_name)
                ws_details.cell(
                    row=row,
                    column=2,
                    value=contract.netsuite_sales_order_number or "",
                )
                ws_details.cell(
                    row=row,
                    column=3,
                    value=contract.netsuite_contract_number or "",
                )
                ws_details.cell(row=row, column=4, value=contract.po_number or "")
                ws_details.cell(row=row, column=5, value=ab_number)
                ws_details.cell(row=row, column=6, value=product_name)
                ws_details.cell(row=row, column=7, value=item.description or "")
                ws_details.cell(
                    row=row, column=8, value=contract.invoice_text or ""
                )
                ws_details.cell(row=row, column=9, value=contract.start_date)
                ws_details.cell(row=row, column=10, value=contract.end_date)
                ws_details.cell(row=row, column=11, value=item.start_date)
                ws_details.cell(row=row, column=12, value=item.billing_start_date)
                ws_details.cell(row=row, column=13, value=item.billing_end_date)
                ws_details.cell(row=row, column=14, value=billing_schedule)
                ws_details.cell(row=row, column=15, value=item.quantity)

                pc = ws_details.cell(
                    row=row, column=16, value=float(item.unit_price)
                )
                pc.number_format = currency_format

                ws_details.cell(
                    row=row,
                    column=17,
                    value=price_period_names.get(
                        item.price_period, item.price_period
                    ),
                )

                tc = ws_details.cell(
                    row=row, column=18, value=float(total_amount)
                )
                tc.number_format = currency_format

                ws_details.cell(
                    row=row,
                    column=19,
                    value="Yes" if item.is_one_off else "",
                )

                row += 1

        detail_widths = [
            35, 18, 30, 15, 15, 35, 30, 40, 14, 14, 14, 14, 14, 18, 10, 15, 15, 15, 10,
        ]
        for col, width in enumerate(detail_widths, 1):
            ws_details.column_dimensions[get_column_letter(col)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
