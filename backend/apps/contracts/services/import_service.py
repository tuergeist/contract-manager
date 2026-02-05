"""Service for importing contracts from NetSuite Excel exports."""

import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import BinaryIO

from django.db import transaction
from openpyxl import load_workbook
from rapidfuzz import fuzz

from apps.contracts.models import Contract, ContractItem
from apps.customers.models import Customer
from apps.products.models import Product


class MatchStatus(Enum):
    """Status of customer matching."""
    MATCHED = "matched"  # High confidence match (>= 0.9)
    REVIEW = "review"    # Needs manual review (< 0.9)
    NOT_FOUND = "not_found"  # No matches found


@dataclass
class MatchResult:
    """Result of customer matching."""
    status: MatchStatus
    customer: Customer | None = None
    confidence: float = 0.0
    alternatives: list["MatchResult"] = field(default_factory=list)
    original_name: str = ""
    netsuite_customer_number: str = ""


@dataclass
class ExcelRow:
    """Parsed row from Excel file."""
    name: str  # Customer name with CUS prefix
    sales_order_number: str
    contract_number: str
    item: str
    invoicing_instructions: str
    renewals_exclusion: str
    contract_start_date: date
    contract_end_date: date
    sum_of_list_rate: Decimal  # Monthly rate
    sum_of_amount: Decimal  # Total amount

    # Extracted fields
    customer_number: str = ""  # e.g., "CUS174"
    customer_name: str = ""    # e.g., "EBARA Corporation"

    def __post_init__(self):
        """Extract customer number and name from the combined name field."""
        if self.name:
            match = re.match(r'^(CUS\d+)\s+(.+)$', self.name)
            if match:
                self.customer_number = match.group(1)
                self.customer_name = match.group(2)


@dataclass
class ContractLineItem:
    """A line item for a contract."""
    item_name: str
    monthly_rate: Decimal
    product: Product | None = None


@dataclass
class ImportProposal:
    """Proposal for a contract import."""
    id: str = field(default_factory=lambda: str(uuid.uuid4()))

    # Source data
    customer_number: str = ""
    customer_name: str = ""
    sales_order_number: str = ""
    contract_number: str = ""
    start_date: date | None = None
    end_date: date | None = None
    invoicing_instructions: str = ""

    # Customer matching
    match_result: MatchResult | None = None
    selected_customer: Customer | None = None

    # Line items
    items: list[ContractLineItem] = field(default_factory=list)

    # Status
    approved: bool = False
    rejected: bool = False
    error: str | None = None

    # Duplicate detection
    existing_contract_id: int | None = None  # If set, this SO already exists

    @property
    def total_monthly_rate(self) -> Decimal:
        """Calculate total monthly rate including discounts (negative line items)."""
        return sum(item.monthly_rate for item in self.items)

    @property
    def needs_review(self) -> bool:
        """Check if this proposal needs manual review."""
        return (
            self.match_result is not None
            and self.match_result.status == MatchStatus.REVIEW
        )


class ExcelParser:
    """Parser for NetSuite Excel exports."""

    # Expected column headers (row 5)
    EXPECTED_HEADERS = [
        "Name",
        "Sales Order Number",
        "Contract (2)",
        "Item",
        "Invoicing Instructions",
        "Renewals Exclusion",
        "Contract Start Date",
        "Contract End Date",
        "Sum of List Rate",
        "Sum of Amount",
    ]

    HEADER_ROW = 5
    DATA_START_ROW = 6

    def __init__(self):
        self.rows: list[ExcelRow] = []
        self.errors: list[str] = []

    def parse(self, file_path: str | Path | BinaryIO) -> list[ExcelRow]:
        """Parse Excel file and return list of rows."""
        self.rows = []
        self.errors = []

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active

        # Validate headers
        headers = [cell.value for cell in sheet[self.HEADER_ROW]]
        for i, expected in enumerate(self.EXPECTED_HEADERS):
            if i >= len(headers) or headers[i] != expected:
                self.errors.append(
                    f"Header mismatch at column {i+1}: "
                    f"expected '{expected}', got '{headers[i] if i < len(headers) else 'missing'}'"
                )

        if self.errors:
            return []

        # Parse data rows
        # Track last values for rows with empty fields (grouped data in Excel)
        last_customer_name: str | None = None
        last_sales_order: str | None = None
        last_contract: str | None = None

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=self.DATA_START_ROW, values_only=True),
            start=self.DATA_START_ROW
        ):
            # Check if row has any meaningful data (item name, dates, or amounts)
            has_data = any(row[i] for i in [3, 6, 7, 8, 9] if i < len(row))

            # Skip completely empty rows
            if not row[0] and not row[1] and not row[2] and not has_data:
                continue

            # Carry forward empty fields from previous rows
            current_name = row[0] if row[0] else last_customer_name
            current_so = row[1] if row[1] else last_sales_order
            current_contract = row[2] if row[2] else last_contract

            # Update tracking when we see new values
            if row[0]:
                last_customer_name = row[0]
            if row[1]:
                last_sales_order = row[1]
            if row[2]:
                last_contract = row[2]

            try:
                excel_row = self._parse_row(
                    row, row_idx,
                    inherited_name=current_name,
                    inherited_so=current_so,
                    inherited_contract=current_contract,
                )
                if excel_row:
                    self.rows.append(excel_row)
            except Exception as e:
                self.errors.append(f"Row {row_idx}: {e}")

        wb.close()
        return self.rows

    def _parse_row(
        self,
        row: tuple,
        row_idx: int,
        inherited_name: str | None = None,
        inherited_so: str | None = None,
        inherited_contract: str | None = None,
    ) -> ExcelRow | None:
        """Parse a single row into an ExcelRow object."""
        # Use inherited values if current row's fields are empty
        name = row[0] if row[0] else inherited_name
        sales_order = row[1] if row[1] else inherited_so
        contract = row[2] if row[2] else inherited_contract

        # Skip rows without customer name
        if not name or not str(name).startswith("CUS"):
            return None

        # Parse dates
        start_date = self._parse_date(row[6])
        end_date = self._parse_date(row[7])

        if not start_date or not end_date:
            self.errors.append(
                f"Row {row_idx}: Invalid dates - start={row[6]}, end={row[7]}"
            )
            return None

        # Parse decimal values
        sum_of_list_rate = self._parse_decimal(row[8])
        sum_of_amount = self._parse_decimal(row[9])

        return ExcelRow(
            name=str(name),
            sales_order_number=str(sales_order or ""),
            contract_number=str(contract or ""),
            item=str(row[3] or ""),
            invoicing_instructions=str(row[4] or ""),
            renewals_exclusion=str(row[5] or ""),
            contract_start_date=start_date,
            contract_end_date=end_date,
            sum_of_list_rate=sum_of_list_rate,
            sum_of_amount=sum_of_amount,
        )

    def _parse_date(self, value) -> date | None:
        """Parse date from Excel cell value."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            # Try common date formats
            for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"]:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def _parse_decimal(self, value) -> Decimal:
        """Parse decimal from Excel cell value."""
        if value is None:
            return Decimal("0")
        if isinstance(value, (int, float)):
            # Round to 2 decimal places
            return Decimal(str(round(value, 2)))
        if isinstance(value, str):
            # Remove currency symbols and whitespace
            cleaned = re.sub(r'[^\d.-]', '', value)
            try:
                return Decimal(cleaned)
            except Exception:
                return Decimal("0")
        return Decimal("0")


class ImportService:
    """Service for importing contracts from parsed Excel data."""

    AUTO_APPROVE_THRESHOLD = 0.9  # Confidence threshold for auto-approval

    def __init__(self, tenant):
        self.tenant = tenant
        self.proposals: list[ImportProposal] = []
        self.errors: list[str] = []

    def generate_proposals(self, rows: list[ExcelRow]) -> list[ImportProposal]:
        """Generate import proposals from parsed Excel rows."""
        self.proposals = []
        self.errors = []

        # Group rows by sales order number
        orders: dict[str, list[ExcelRow]] = {}
        for row in rows:
            key = row.sales_order_number
            if key not in orders:
                orders[key] = []
            orders[key].append(row)

        # Generate a proposal for each order
        for _sales_order, order_rows in orders.items():
            proposal = self._create_proposal(order_rows)
            if proposal:
                self.proposals.append(proposal)

        return self.proposals

    def _create_proposal(self, rows: list[ExcelRow]) -> ImportProposal | None:
        """Create a proposal from a group of Excel rows for the same order."""
        if not rows:
            return None

        # All rows should have the same customer and order info
        first = rows[0]

        proposal = ImportProposal(
            customer_number=first.customer_number,
            customer_name=first.customer_name,
            sales_order_number=first.sales_order_number,
            contract_number=first.contract_number,
            start_date=first.contract_start_date,
            end_date=first.contract_end_date,
            invoicing_instructions=first.invoicing_instructions,
        )

        # Check if this contract already exists
        existing = Contract.objects.filter(
            tenant=self.tenant,
            netsuite_sales_order_number=first.sales_order_number,
        ).first()
        if existing:
            proposal.existing_contract_id = existing.id

        # Match customer
        proposal.match_result = self._match_customer(
            first.customer_name,
            first.customer_number,
        )

        if proposal.match_result.status == MatchStatus.MATCHED:
            proposal.selected_customer = proposal.match_result.customer

        # Add line items
        for row in rows:
            if row.item == "Sales Discount":
                # Add discount as a line item with negative rate
                item = ContractLineItem(
                    item_name=row.item,
                    monthly_rate=row.sum_of_list_rate,
                    product=None,
                )
            else:
                # Regular item
                item = ContractLineItem(
                    item_name=row.item,
                    monthly_rate=row.sum_of_list_rate,
                    product=self._match_product(row.item),
                )
            proposal.items.append(item)

        return proposal

    def _match_customer(
        self,
        name: str,
        netsuite_number: str,
    ) -> MatchResult:
        """Match customer name against existing customers."""
        # First try to find by NetSuite customer number
        if netsuite_number:
            customer = Customer.objects.filter(
                tenant=self.tenant,
                netsuite_customer_number=netsuite_number,
            ).first()
            if customer:
                return MatchResult(
                    status=MatchStatus.MATCHED,
                    customer=customer,
                    confidence=1.0,
                    original_name=name,
                    netsuite_customer_number=netsuite_number,
                )

        # Fuzzy match against customer names (match all customers, not just active)
        customers = Customer.objects.filter(tenant=self.tenant)
        matches = []

        for customer in customers:
            # Use weighted ratio for stricter matching
            # token_set_ratio was too permissive (matching "International" in unrelated companies)
            score = fuzz.WRatio(name.lower(), customer.name.lower()) / 100.0
            if score > 0.75:  # Only consider matches above 75%
                matches.append(MatchResult(
                    status=MatchStatus.MATCHED if score >= self.AUTO_APPROVE_THRESHOLD else MatchStatus.REVIEW,
                    customer=customer,
                    confidence=score,
                    original_name=name,
                    netsuite_customer_number=netsuite_number,
                ))

        # Sort by confidence
        matches.sort(key=lambda x: x.confidence, reverse=True)

        if not matches:
            return MatchResult(
                status=MatchStatus.NOT_FOUND,
                original_name=name,
                netsuite_customer_number=netsuite_number,
            )

        best_match = matches[0]
        if best_match.confidence >= self.AUTO_APPROVE_THRESHOLD:
            best_match.status = MatchStatus.MATCHED
        else:
            best_match.status = MatchStatus.REVIEW
            best_match.alternatives = matches[1:4]  # Include up to 3 alternatives

        return best_match

    def _match_product(self, item_name: str) -> Product | None:
        """Match item name against existing products."""
        # First try exact match by netsuite_item_name
        product = Product.objects.filter(
            tenant=self.tenant,
            netsuite_item_name=item_name,
        ).first()
        if product:
            return product

        # Try fuzzy match against product names
        products = Product.objects.filter(tenant=self.tenant, is_active=True)
        best_match = None
        best_score = 0.0

        for product in products:
            score = fuzz.token_set_ratio(
                item_name.lower(),
                product.name.lower()
            ) / 100.0
            if score > best_score and score > 0.7:
                best_match = product
                best_score = score

        return best_match

    def apply_proposals(
        self,
        proposals: list[ImportProposal] | None = None,
        auto_create_products: bool = True,
    ) -> list[Contract]:
        """Apply approved proposals and create contracts."""
        if proposals is None:
            proposals = [p for p in self.proposals if p.approved and not p.rejected]

        created_contracts = []

        for proposal in proposals:
            if proposal.rejected:
                continue
            if not proposal.approved:
                continue
            if not proposal.selected_customer:
                proposal.error = "No customer selected"
                continue

            # Check for duplicate contract
            existing = Contract.objects.filter(
                tenant=self.tenant,
                netsuite_sales_order_number=proposal.sales_order_number,
            ).first()
            if existing:
                proposal.error = f"Contract with SO {proposal.sales_order_number} already exists (ID: {existing.id})"
                continue

            try:
                contract = self._create_contract(proposal, auto_create_products)
                created_contracts.append(contract)
            except Exception as e:
                proposal.error = str(e)

        return created_contracts

    @transaction.atomic
    def _create_contract(
        self,
        proposal: ImportProposal,
        auto_create_products: bool,
    ) -> Contract:
        """Create a contract from a proposal."""
        # Update customer with NetSuite number if not set
        customer = proposal.selected_customer
        if proposal.customer_number and not customer.netsuite_customer_number:
            customer.netsuite_customer_number = proposal.customer_number
            customer.save(update_fields=["netsuite_customer_number"])

        # Create contract
        contract = Contract.objects.create(
            tenant=self.tenant,
            customer=customer,
            name=proposal.sales_order_number,
            status=Contract.Status.DRAFT,
            start_date=proposal.start_date,
            end_date=proposal.end_date,
            billing_start_date=proposal.start_date,
            billing_interval=Contract.BillingInterval.ANNUAL,  # NetSuite contracts are typically annual
            netsuite_sales_order_number=proposal.sales_order_number,
            netsuite_contract_number=proposal.contract_number,
            notes=proposal.invoicing_instructions,
        )

        # Create contract items
        for item in proposal.items:
            product = item.product

            if item.monthly_rate < 0:
                # Discount line item (no product)
                ContractItem.objects.create(
                    tenant=self.tenant,
                    contract=contract,
                    product=None,
                    description=item.item_name,
                    quantity=1,
                    unit_price=item.monthly_rate,
                    price_source=ContractItem.PriceSource.CUSTOM,
                )
                continue

            # Auto-create product if needed
            if not product and auto_create_products:
                product = Product.objects.create(
                    tenant=self.tenant,
                    name=item.item_name,
                    netsuite_item_name=item.item_name,
                    type=Product.ProductType.SUBSCRIPTION,
                    is_active=True,
                )

            if product:
                ContractItem.objects.create(
                    tenant=self.tenant,
                    contract=contract,
                    product=product,
                    quantity=1,
                    unit_price=item.monthly_rate,
                    price_source=ContractItem.PriceSource.CUSTOM,
                )

        return contract

    def get_summary(self) -> dict:
        """Get a summary of the import proposals."""
        total = len(self.proposals)
        matched = sum(
            1 for p in self.proposals
            if p.match_result and p.match_result.status == MatchStatus.MATCHED
        )
        needs_review = sum(
            1 for p in self.proposals
            if p.match_result and p.match_result.status == MatchStatus.REVIEW
        )
        not_found = sum(
            1 for p in self.proposals
            if p.match_result and p.match_result.status == MatchStatus.NOT_FOUND
        )

        return {
            "total_proposals": total,
            "auto_matched": matched,
            "needs_review": needs_review,
            "not_found": not_found,
            "total_items": sum(len(p.items) for p in self.proposals),
            "errors": self.errors,
        }
