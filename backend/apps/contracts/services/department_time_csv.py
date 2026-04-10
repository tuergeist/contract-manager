"""Generate department time analysis as Excel file for email attachment."""
import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, numbers


def generate_department_time_xlsx(tenant, year: int, month: int) -> tuple[bytes, str]:
    """Generate XLSX for department time analysis. Returns (xlsx_bytes, filename)."""
    from dateutil.relativedelta import relativedelta

    from apps.core.context import Context
    from config.schema import schema

    date_from = date(year, month, 1)
    date_to = (date_from + relativedelta(months=1)) - timedelta(days=1)

    admin_user = tenant.users.filter(is_admin=True).first()
    if not admin_user:
        admin_user = tenant.users.first()

    class FakeRequest:
        headers = {}

    ctx = Context(request=FakeRequest(), user=admin_user)

    result = schema.execute_sync(
        """
        query($dateFrom: Date!, $dateTo: Date!) {
            departmentTimeAnalysis(dateFrom: $dateFrom, dateTo: $dateTo) {
                costDistribution {
                    departmentName
                    ftes
                    percentage
                    cost
                }
            }
        }
        """,
        variable_values={"dateFrom": str(date_from), "dateTo": str(date_to)},
        context_value=ctx,
    )

    if result.errors:
        raise ValueError(f"Query failed: {result.errors}")

    rows = result.data["departmentTimeAnalysis"]["costDistribution"] or []

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    # Header
    headers = ["Department", "FTEs", "%"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10

    # Data
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["departmentName"])
        ws.cell(row=i, column=2, value=round(r["ftes"], 2))
        ws.cell(row=i, column=3, value=round(r["percentage"], 1))

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"department-time-{year}-{month:02d}.xlsx"
    return buf.getvalue(), filename
