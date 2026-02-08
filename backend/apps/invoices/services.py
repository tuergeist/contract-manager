"""Invoice service for generating and exporting invoices."""
import io
import zipfile
from calendar import monthrange
from datetime import date
from decimal import Decimal
from typing import Literal

from django.db import transaction
from django.template.loader import render_to_string
from django.utils import timezone

try:
    from weasyprint import HTML
except ImportError:
    HTML = None  # PDF generation will be unavailable

from apps.contracts.models import Contract
from apps.invoices.types import InvoiceData, InvoiceLineItem
from apps.tenants.models import Tenant


# Localization labels for invoice PDF
LABELS = {
    "de": {
        "invoice": "Rechnung",
        "invoice_no": "Rechnungsnr.",
        "bill_to": "Rechnungsadresse",
        "invoice_date": "Rechnungsdatum",
        "billing_period": "Abrechnungszeitraum",
        "service_period": "Leistungszeitraum",
        "contract": "Vertrag",
        "description": "Beschreibung",
        "quantity": "Menge",
        "unit_price": "Einzelpreis",
        "amount": "Betrag",
        "net_total": "Nettobetrag",
        "tax": "MwSt.",
        "gross_total": "Bruttobetrag",
        "total": "Gesamtbetrag",
        "prorated": "Anteilig",
        "one_off": "Einmalig",
        "customer_id": "Kunden-Nr.",
        "vat_id": "USt-IdNr.",
        "tax_number": "Steuernummer",
        "register": "Handelsregister",
        "managing_directors": "Geschäftsführer",
        "share_capital": "Stammkapital",
        "bank_details": "Bankverbindung",
        "phone": "Telefon",
        "pos": "Pos.",
        "date_label": "Datum",
        "invoice_amount": "Rechnungsbetrag",
    },
    "en": {
        "invoice": "Invoice",
        "invoice_no": "Invoice No.",
        "bill_to": "Bill To",
        "invoice_date": "Invoice Date",
        "billing_period": "Billing Period",
        "service_period": "Service Period",
        "contract": "Contract",
        "description": "Description",
        "quantity": "Qty",
        "unit_price": "Unit Price",
        "amount": "Amount",
        "net_total": "Net Total",
        "tax": "VAT",
        "gross_total": "Gross Total",
        "total": "Total",
        "prorated": "Prorated",
        "one_off": "One-time",
        "customer_id": "Customer ID",
        "vat_id": "VAT ID",
        "tax_number": "Tax Number",
        "register": "Commercial Register",
        "managing_directors": "Managing Directors",
        "share_capital": "Share Capital",
        "bank_details": "Bank Details",
        "phone": "Phone",
        "pos": "Pos.",
        "date_label": "Date",
        "invoice_amount": "Invoice Total",
    },
}


class InvoiceService:
    """Service for generating invoices from contract billing schedules."""

    def __init__(self, tenant: Tenant):
        self.tenant = tenant

    def get_invoices_for_month(self, year: int, month: int) -> list[InvoiceData]:
        """
        Get all invoices due for a specific month.

        Aggregates billing events from all active contracts within the tenant
        for the specified month.

        Args:
            year: The year (e.g., 2026)
            month: The month (1-12)

        Returns:
            List of InvoiceData objects, one per contract with billing events
            in the specified month.
        """
        # Calculate date range for the month
        first_day = date(year, month, 1)
        last_day = date(year, month, monthrange(year, month)[1])

        # Get all active contracts for this tenant
        contracts = Contract.objects.filter(
            tenant=self.tenant,
            status=Contract.Status.ACTIVE,
        ).select_related("customer").prefetch_related("items__product")

        invoices = []

        for contract in contracts:
            # Get billing schedule for this contract within the month
            billing_events = contract.get_billing_schedule(
                from_date=first_day,
                to_date=last_day,
            )

            # Build items lookup once per contract (uses prefetched data, avoids N+1)
            contract_items_by_id = {ci.id: ci for ci in contract.items.all()}

            # Create an invoice for each billing event in this month
            for event in billing_events:
                billing_date = event["date"]

                # Skip if billing date is outside our target month
                if billing_date.year != year or billing_date.month != month:
                    continue

                # Skip if no items (shouldn't happen, but defensive)
                if not event["items"]:
                    continue

                # Calculate billing period (depends on contract interval)
                period_start, period_end = self._calculate_billing_period(
                    contract, billing_date
                )

                # Convert billing items to invoice line items
                line_items = []
                for item in event["items"]:
                    # Get the actual ContractItem for additional fields
                    contract_item = contract_items_by_id.get(item["item_id"])

                    line_item = InvoiceLineItem(
                        item_id=item["item_id"],
                        product_name=item["product_name"],
                        description=item.get("description", ""),
                        quantity=item["quantity"],
                        unit_price=item["unit_price"],
                        amount=item["amount"],
                        is_prorated=item.get("is_prorated", False),
                        prorate_factor=item.get("prorate_factor"),
                        is_one_off=item.get("is_one_off", False),
                        # Item-level fields
                        item_start_date=contract_item.start_date if contract_item else None,
                        item_billing_start_date=contract_item.billing_start_date if contract_item else None,
                        item_billing_end_date=contract_item.billing_end_date if contract_item else None,
                        order_confirmation_number=contract_item.order_confirmation_number if contract_item else None,
                    )
                    line_items.append(line_item)

                invoice = InvoiceData(
                    contract_id=contract.id,
                    contract_name=contract.name or f"Contract {contract.id}",
                    customer_id=contract.customer.id,
                    customer_name=contract.customer.name,
                    customer_address=contract.customer.address or {},
                    billing_date=billing_date,
                    billing_period_start=period_start,
                    billing_period_end=period_end,
                    line_items=line_items,
                    invoice_text=contract.invoice_text or "",
                    # Enhanced fields
                    customer_number=contract.customer.netsuite_customer_number or "",
                    sales_order_number=contract.netsuite_sales_order_number or "",
                    contract_number=contract.netsuite_contract_number or "",
                    po_number=contract.po_number or "",
                    order_confirmation_number=contract.order_confirmation_number or "",
                    contract_start_date=contract.start_date,
                    contract_end_date=contract.end_date,
                    billing_interval=contract.billing_interval,
                )
                invoices.append(invoice)

        # Sort by customer name for consistent ordering
        invoices.sort(key=lambda x: x.customer_name.lower())

        return invoices

    def _calculate_billing_period(
        self, contract: Contract, billing_date: date
    ) -> tuple[date, date]:
        """Calculate the billing period for a given billing date."""
        from dateutil.relativedelta import relativedelta

        interval_months = contract.get_interval_months()

        # Period starts at billing_date
        period_start = billing_date

        # Period ends before next billing date
        period_end = billing_date + relativedelta(months=interval_months, days=-1)

        # Don't extend beyond contract end date
        if contract.end_date and period_end > contract.end_date:
            period_end = contract.end_date

        return period_start, period_end

    def _get_template_context(self) -> dict:
        """Load template settings and legal data for PDF rendering.

        Returns a dict with company, accent_color, header_text, footer_text,
        logo_url, and tax_rate. Falls back to defaults when not configured.
        """
        from apps.invoices.models import CompanyLegalData, InvoiceTemplate

        # Company legal data (fallback to tenant name)
        try:
            legal_data = self.tenant.legal_data
            company = legal_data.to_snapshot()
        except CompanyLegalData.DoesNotExist:
            company = {
                "company_name": self.tenant.name,
                "street": "",
                "zip_code": "",
                "city": "",
                "country": "",
                "tax_number": "",
                "vat_id": "",
                "commercial_register_court": "",
                "commercial_register_number": "",
                "managing_directors": [],
                "bank_name": "",
                "iban": "",
                "bic": "",
                "phone": "",
                "email": "",
                "website": "",
                "share_capital": "",
                "default_tax_rate": "19.00",
            }

        # Template settings (fallback to defaults)
        accent_color = "#2563eb"
        header_text = ""
        footer_text = ""
        logo_url = ""
        try:
            template = InvoiceTemplate.objects.get(tenant=self.tenant)
            accent_color = template.accent_color or "#2563eb"
            header_text = template.header_text or ""
            footer_text = template.footer_text or ""
            if template.logo and template.logo.name:
                # Use file:// URI so WeasyPrint can resolve the logo in PDF rendering
                import base64
                import mimetypes
                try:
                    mime_type = mimetypes.guess_type(template.logo.name)[0] or "image/png"
                    logo_data = template.logo.read()
                    logo_url = f"data:{mime_type};base64,{base64.b64encode(logo_data).decode()}"
                except Exception:
                    logo_url = ""
        except InvoiceTemplate.DoesNotExist:
            pass

        tax_rate = Decimal(company.get("default_tax_rate", "19.00"))

        return {
            "company": company,
            "accent_color": accent_color,
            "header_text": header_text,
            "footer_text": footer_text,
            "logo_url": logo_url,
            "tax_rate": tax_rate,
        }

    def generate_pdf(
        self,
        invoices: list[InvoiceData],
        language: Literal["de", "en"] = "de",
    ) -> bytes:
        """
        Generate a combined PDF containing all invoices.

        Each invoice starts on a new page. Includes company legal data,
        tax breakdown, template customization, and GmbH legal footer.

        Args:
            invoices: List of InvoiceData to include in PDF
            language: Language for labels ("de" or "en")

        Returns:
            PDF file as bytes.
        """
        if not invoices:
            return b""

        labels = LABELS.get(language, LABELS["en"])
        currency_symbol = self.tenant.currency_symbol
        template_ctx = self._get_template_context()
        tax_rate = template_ctx["tax_rate"]

        # Render each invoice as HTML
        html_parts = []
        for i, invoice in enumerate(invoices):
            # Calculate tax for this invoice
            total_net = invoice.total_amount
            tax_amount, total_gross = self.calculate_tax(total_net, tax_rate)

            # Convert dataclass to dict for template
            invoice_dict = {
                "contract_id": invoice.contract_id,
                "contract_name": invoice.contract_name,
                "customer_id": invoice.customer_id,
                "customer_name": invoice.customer_name,
                "customer_address": invoice.customer_address,
                "billing_date": invoice.billing_date,
                "billing_period_start": invoice.billing_period_start,
                "billing_period_end": invoice.billing_period_end,
                "line_items": [
                    {
                        "item_id": item.item_id,
                        "product_name": item.product_name,
                        "description": item.description,
                        "quantity": item.quantity,
                        "unit_price": item.unit_price,
                        "amount": item.amount,
                        "is_prorated": item.is_prorated,
                        "prorate_factor": item.prorate_factor,
                        "is_one_off": item.is_one_off,
                    }
                    for item in invoice.line_items
                ],
                "total_amount": invoice.total_amount,
                "total_net": total_net,
                "tax_amount": tax_amount,
                "total_gross": total_gross,
            }

            html = render_to_string(
                "invoices/invoice.html",
                {
                    "invoice": invoice_dict,
                    "labels": labels,
                    "language": language,
                    "currency_symbol": currency_symbol,
                    "invoice_number": getattr(invoice, "invoice_number", ""),
                    "tax_rate": tax_rate,
                    **template_ctx,
                },
            )
            html_parts.append(html)

        # Combine all HTML parts with page breaks
        combined_html = '<div class="page-break"></div>'.join(html_parts)

        # Generate PDF
        pdf_document = HTML(string=combined_html).render()
        return pdf_document.write_pdf()

    def generate_preview_pdf(
        self,
        language: Literal["de", "en"] = "de",
    ) -> bytes:
        """Generate a sample invoice PDF using current template settings and dummy data."""
        labels = LABELS.get(language, LABELS["en"])
        currency_symbol = self.tenant.currency_symbol
        template_ctx = self._get_template_context()
        tax_rate = template_ctx["tax_rate"]

        today = date.today()
        net = Decimal("1500.00")
        tax_amount, gross = self.calculate_tax(net, tax_rate)

        invoice_dict = {
            "contract_name": "Mustervertrag 2025-001",
            "customer_name": "Mustermann GmbH",
            "customer_address": {
                "street": "Musterstraße 1",
                "zip_code": "12345",
                "city": "Musterstadt",
                "country": "Deutschland",
            },
            "billing_date": today,
            "billing_period_start": today.replace(day=1),
            "billing_period_end": today,
            "line_items": [
                {
                    "product_name": "Software-Lizenz Premium",
                    "description": "Monatliche Lizenzgebühr",
                    "quantity": 5,
                    "unit_price": Decimal("200.00"),
                    "amount": Decimal("1000.00"),
                    "is_prorated": False,
                    "prorate_factor": None,
                    "is_one_off": False,
                },
                {
                    "product_name": "Support & Wartung",
                    "description": "Wartungsvertrag",
                    "quantity": 1,
                    "unit_price": Decimal("500.00"),
                    "amount": Decimal("500.00"),
                    "is_prorated": False,
                    "prorate_factor": None,
                    "is_one_off": False,
                },
            ],
            "total_amount": net,
            "total_net": net,
            "tax_amount": tax_amount,
            "total_gross": gross,
        }

        html = render_to_string(
            "invoices/invoice.html",
            {
                "invoice": invoice_dict,
                "labels": labels,
                "language": language,
                "currency_symbol": currency_symbol,
                "invoice_number": "PREVIEW-2025-0001",
                "tax_rate": tax_rate,
                **template_ctx,
            },
        )

        pdf_document = HTML(string=html).render()
        return pdf_document.write_pdf()

    def generate_individual_pdfs(
        self,
        invoices: list[InvoiceData],
        year: int,
        month: int,
        language: Literal["de", "en"] = "de",
    ) -> bytes:
        """
        Generate individual PDFs for each invoice, packaged as a ZIP.

        Args:
            invoices: List of InvoiceData to generate PDFs for
            year: Year for filename
            month: Month for filename
            language: Language for labels ("de" or "en")

        Returns:
            ZIP file containing individual PDFs as bytes.
        """
        if not invoices:
            return b""

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for invoice in invoices:
                # Generate PDF for single invoice
                pdf_bytes = self.generate_pdf([invoice], language)

                # Create safe filename
                customer_safe = self._safe_filename(invoice.customer_name)
                contract_safe = self._safe_filename(invoice.contract_name)
                filename = f"invoice-{customer_safe}-{contract_safe}-{year:04d}-{month:02d}.pdf"

                zip_file.writestr(filename, pdf_bytes)

        return zip_buffer.getvalue()

    def _safe_filename(self, name: str) -> str:
        """Convert a name to a safe filename component."""
        import re
        # Replace spaces with hyphens, remove special characters
        safe = re.sub(r"[^\w\s-]", "", name)
        safe = re.sub(r"[-\s]+", "-", safe).strip("-")
        return safe[:50]  # Limit length

    def generate_excel(
        self,
        invoices: list[InvoiceData],
        year: int,
        month: int,
        language: Literal["de", "en"] = "de",
    ) -> bytes:
        """
        Generate an Excel file with all invoices in NetSuite-compatible format.

        Creates two sheets:
        - Summary: Pivot-style overview grouped by customer/contract
        - Details: Full line item details with all metadata

        Args:
            invoices: List of InvoiceData to include
            year: Year for context
            month: Month for context
            language: Language for headers ("de" or "en")

        Returns:
            Excel file as bytes.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not invoices:
            # Return empty workbook
            wb = Workbook()
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        # Localized headers
        headers_de = {
            "customer": "Kunde",
            "sales_order": "SO-Nummer",
            "contract_number": "Vertrag",
            "po_number": "Bestellnummer",
            "ab_number": "AB-Nummer",
            "item": "Position",
            "item_description": "Beschreibung",
            "invoicing_instructions": "Rechnungshinweise",
            "contract_start": "Vertragsbeginn",
            "contract_end": "Vertragsende",
            "item_start": "Position gültig ab",
            "item_billing_start": "Abrechnung ab",
            "item_billing_end": "Abrechnung bis",
            "billing_schedule": "Abrechnungsintervall",
            "quantity": "Menge",
            "unit_price": "Einzelpreis",
            "amount": "Betrag",
            "billing_date": "Rechnungsdatum",
            "period_start": "Zeitraum von",
            "period_end": "Zeitraum bis",
            "is_prorated": "Anteilig",
            "is_one_off": "Einmalig",
            "total": "Gesamtbetrag",
            "monthly_rate": "Monatliche Rate",
            "summary_title": "Rechnungsübersicht",
        }
        headers_en = {
            "customer": "Customer",
            "sales_order": "Sales Order",
            "contract_number": "Contract",
            "po_number": "PO Number",
            "ab_number": "Order Confirmation",
            "item": "Item",
            "item_description": "Description",
            "invoicing_instructions": "Invoicing Instructions",
            "contract_start": "Contract Start",
            "contract_end": "Contract End",
            "item_start": "Item Effective Date",
            "item_billing_start": "Billing Start",
            "item_billing_end": "Billing End",
            "billing_schedule": "Billing Schedule",
            "quantity": "Quantity",
            "unit_price": "Unit Price",
            "amount": "Amount",
            "billing_date": "Billing Date",
            "period_start": "Period Start",
            "period_end": "Period End",
            "is_prorated": "Prorated",
            "is_one_off": "One-off",
            "total": "Total",
            "monthly_rate": "Monthly Rate",
            "summary_title": "Invoice Summary",
        }
        h = headers_de if language == "de" else headers_en

        # Billing interval display names
        interval_names = {
            "monthly": "Monthly" if language == "en" else "Monatlich",
            "quarterly": "Quarterly" if language == "en" else "Vierteljährlich",
            "semi_annual": "Semi-annual" if language == "en" else "Halbjährlich",
            "annual": "Annual" if language == "en" else "Jährlich",
            "biennial": "2 Years" if language == "en" else "2 Jahre",
            "triennial": "3 Years" if language == "en" else "3 Jahre",
            "quadrennial": "4 Years" if language == "en" else "4 Jahre",
            "quinquennial": "5 Years" if language == "en" else "5 Jahre",
        }

        wb = Workbook()
        currency = self.tenant.currency
        currency_format = f'#,##0.00 "{currency}"'

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        total_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # =================================================================
        # Summary Sheet - Pivot-style overview
        # =================================================================
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Title
        ws_summary["A1"] = f"{h['summary_title']} - {year:04d}-{month:02d}"
        ws_summary["A1"].font = title_font

        # Summary headers
        summary_headers = [
            h["customer"],
            h["sales_order"],
            h["contract_number"],
            h["po_number"],
            h["invoicing_instructions"],
            h["contract_start"],
            h["contract_end"],
            h["billing_schedule"],
            h["monthly_rate"],
            h["amount"],
        ]

        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Aggregate data by contract
        contract_data: dict[int, dict] = {}
        for inv in invoices:
            if inv.contract_id not in contract_data:
                # Calculate monthly rate (amount / interval months)
                interval_months = {
                    "monthly": 1, "quarterly": 3, "semi_annual": 6,
                    "annual": 12, "biennial": 24, "triennial": 36,
                    "quadrennial": 48, "quinquennial": 60,
                }.get(inv.billing_interval, 1)
                monthly_rate = inv.total_amount / Decimal(interval_months)

                contract_data[inv.contract_id] = {
                    "customer": inv.customer_display_name,
                    "sales_order": inv.sales_order_number,
                    "contract_number": inv.contract_number,
                    "po_number": inv.po_number,
                    "invoicing_instructions": inv.invoice_text,
                    "contract_start": inv.contract_start_date,
                    "contract_end": inv.contract_end_date,
                    "billing_schedule": interval_names.get(inv.billing_interval, inv.billing_interval),
                    "monthly_rate": monthly_rate,
                    "amount": inv.total_amount,
                }
            else:
                # Accumulate amounts for same contract
                contract_data[inv.contract_id]["amount"] += inv.total_amount

        # Write summary rows
        row = 4
        total_amount = Decimal("0")
        total_monthly = Decimal("0")
        for contract_id, data in sorted(contract_data.items(), key=lambda x: x[1]["customer"].lower()):
            ws_summary.cell(row=row, column=1, value=data["customer"])
            ws_summary.cell(row=row, column=2, value=data["sales_order"])
            ws_summary.cell(row=row, column=3, value=data["contract_number"])
            ws_summary.cell(row=row, column=4, value=data["po_number"])
            ws_summary.cell(row=row, column=5, value=data["invoicing_instructions"])
            ws_summary.cell(row=row, column=6, value=data["contract_start"])
            ws_summary.cell(row=row, column=7, value=data["contract_end"])
            ws_summary.cell(row=row, column=8, value=data["billing_schedule"])

            monthly_cell = ws_summary.cell(row=row, column=9, value=float(data["monthly_rate"]))
            monthly_cell.number_format = currency_format

            amount_cell = ws_summary.cell(row=row, column=10, value=float(data["amount"]))
            amount_cell.number_format = currency_format

            total_amount += data["amount"]
            total_monthly += data["monthly_rate"]
            row += 1

        # Total row
        ws_summary.cell(row=row, column=1, value=h["total"]).font = Font(bold=True)
        for col in range(1, 9):
            ws_summary.cell(row=row, column=col).fill = total_fill

        monthly_total_cell = ws_summary.cell(row=row, column=9, value=float(total_monthly))
        monthly_total_cell.number_format = currency_format
        monthly_total_cell.font = Font(bold=True)
        monthly_total_cell.fill = total_fill

        total_cell = ws_summary.cell(row=row, column=10, value=float(total_amount))
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True)
        total_cell.fill = total_fill

        # Adjust column widths for Summary
        summary_widths = [35, 18, 30, 15, 40, 14, 14, 18, 15, 15]
        for col, width in enumerate(summary_widths, 1):
            ws_summary.column_dimensions[get_column_letter(col)].width = width

        # =================================================================
        # Details Sheet - Full line item data
        # =================================================================
        ws_details = wb.create_sheet("Details")

        detail_headers = [
            h["customer"],
            h["sales_order"],
            h["contract_number"],
            h["po_number"],
            h["ab_number"],
            h["item"],
            h["item_description"],
            h["invoicing_instructions"],
            h["contract_start"],
            h["contract_end"],
            h["item_start"],
            h["item_billing_start"],
            h["item_billing_end"],
            h["billing_schedule"],
            h["billing_date"],
            h["period_start"],
            h["period_end"],
            h["quantity"],
            h["unit_price"],
            h["amount"],
            h["is_prorated"],
            h["is_one_off"],
        ]

        for col, header in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Write detail rows
        row = 2
        for inv in invoices:
            billing_schedule = interval_names.get(inv.billing_interval, inv.billing_interval)

            for item in inv.line_items:
                # Use item-level AB number if available, otherwise contract-level
                ab_number = item.order_confirmation_number or inv.order_confirmation_number

                ws_details.cell(row=row, column=1, value=inv.customer_display_name)
                ws_details.cell(row=row, column=2, value=inv.sales_order_number)
                ws_details.cell(row=row, column=3, value=inv.contract_number)
                ws_details.cell(row=row, column=4, value=inv.po_number)
                ws_details.cell(row=row, column=5, value=ab_number)
                ws_details.cell(row=row, column=6, value=item.product_name)
                ws_details.cell(row=row, column=7, value=item.description)
                ws_details.cell(row=row, column=8, value=inv.invoice_text)
                ws_details.cell(row=row, column=9, value=inv.contract_start_date)
                ws_details.cell(row=row, column=10, value=inv.contract_end_date)
                ws_details.cell(row=row, column=11, value=item.item_start_date)
                ws_details.cell(row=row, column=12, value=item.item_billing_start_date)
                ws_details.cell(row=row, column=13, value=item.item_billing_end_date)
                ws_details.cell(row=row, column=14, value=billing_schedule)
                ws_details.cell(row=row, column=15, value=inv.billing_date)
                ws_details.cell(row=row, column=16, value=inv.billing_period_start)
                ws_details.cell(row=row, column=17, value=inv.billing_period_end)
                ws_details.cell(row=row, column=18, value=item.quantity)

                price_cell = ws_details.cell(row=row, column=19, value=float(item.unit_price))
                price_cell.number_format = currency_format

                amount_cell = ws_details.cell(row=row, column=20, value=float(item.amount))
                amount_cell.number_format = currency_format

                ws_details.cell(row=row, column=21, value="Yes" if item.is_prorated else "")
                ws_details.cell(row=row, column=22, value="Yes" if item.is_one_off else "")

                row += 1

        # Adjust column widths for Details
        detail_widths = [35, 18, 30, 15, 15, 35, 30, 40, 14, 14, 14, 14, 14, 18, 14, 14, 14, 10, 15, 15, 10, 10]
        for col, width in enumerate(detail_widths, 1):
            ws_details.column_dimensions[get_column_letter(col)].width = width

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ----------------------------------------------------------------
    # Tax calculation
    # ----------------------------------------------------------------

    def get_tax_rate(self) -> Decimal:
        """Get the tenant's default tax rate."""
        try:
            return self.tenant.legal_data.default_tax_rate
        except Exception:
            return Decimal("19.00")

    @staticmethod
    def calculate_tax(
        net_amount: Decimal, tax_rate: Decimal
    ) -> tuple[Decimal, Decimal]:
        """Calculate tax amount and gross from net and rate.

        Returns (tax_amount, gross_amount).
        """
        tax_amount = (net_amount * tax_rate / Decimal("100")).quantize(Decimal("0.01"))
        gross_amount = net_amount + tax_amount
        return tax_amount, gross_amount

    # ----------------------------------------------------------------
    # Invoice persistence
    # ----------------------------------------------------------------

    def generate_and_persist(self, year: int, month: int) -> list:
        """
        Generate invoices for a month, assign numbers, and persist as records.

        Returns list of created InvoiceRecord instances.
        Skips contracts that already have a finalized invoice for the same period.
        """
        from apps.invoices.models import CompanyLegalData, InvoiceRecord
        from apps.invoices.numbering import InvoiceNumberService

        # Validate legal data exists
        try:
            legal_data = self.tenant.legal_data
        except CompanyLegalData.DoesNotExist:
            raise ValueError(
                "Company legal data must be configured before generating invoices."
            )

        # Calculate invoices
        invoices = self.get_invoices_for_month(year, month)
        if not invoices:
            return []

        tax_rate = legal_data.default_tax_rate
        company_snapshot = legal_data.to_snapshot()
        numbering = InvoiceNumberService(self.tenant)

        created_records = []

        with transaction.atomic():
            for invoice_data in invoices:
                # Check for existing finalized invoice
                exists = InvoiceRecord.objects.filter(
                    tenant=self.tenant,
                    contract_id=invoice_data.contract_id,
                    billing_date=invoice_data.billing_date,
                    period_start=invoice_data.billing_period_start,
                    period_end=invoice_data.billing_period_end,
                    status__in=[
                        InvoiceRecord.Status.FINALIZED,
                        InvoiceRecord.Status.DRAFT,
                    ],
                ).exists()
                if exists:
                    continue

                # Calculate amounts
                total_net = invoice_data.total_amount
                tax_amount, total_gross = self.calculate_tax(total_net, tax_rate)

                # Snapshot line items
                line_items_snapshot = [
                    {
                        "item_id": item.item_id,
                        "product_name": item.product_name,
                        "description": item.description,
                        "quantity": item.quantity,
                        "unit_price": str(item.unit_price),
                        "amount": str(item.amount),
                        "is_prorated": item.is_prorated,
                        "prorate_factor": str(item.prorate_factor) if item.prorate_factor else None,
                        "is_one_off": item.is_one_off,
                    }
                    for item in invoice_data.line_items
                ]

                # Get next invoice number
                invoice_number = numbering.get_next_number(invoice_data.billing_date)

                record = InvoiceRecord.objects.create(
                    tenant=self.tenant,
                    contract_id=invoice_data.contract_id,
                    customer_id=invoice_data.customer_id,
                    invoice_number=invoice_number,
                    billing_date=invoice_data.billing_date,
                    period_start=invoice_data.billing_period_start,
                    period_end=invoice_data.billing_period_end,
                    total_net=total_net,
                    tax_rate=tax_rate,
                    tax_amount=tax_amount,
                    total_gross=total_gross,
                    line_items_snapshot=line_items_snapshot,
                    company_data_snapshot=company_snapshot,
                    status=InvoiceRecord.Status.FINALIZED,
                    customer_name=invoice_data.customer_name,
                    contract_name=invoice_data.contract_name,
                    invoice_text=invoice_data.invoice_text,
                )
                created_records.append(record)

        return created_records

    @staticmethod
    def cancel_invoice(invoice_record) -> None:
        """Cancel a finalized invoice. Number is NOT reused."""
        from apps.invoices.models import InvoiceRecord

        if invoice_record.status != InvoiceRecord.Status.FINALIZED:
            raise ValueError("Only finalized invoices can be cancelled.")
        invoice_record.status = InvoiceRecord.Status.CANCELLED
        invoice_record.save(update_fields=["status", "updated_at"])

    def get_persisted_invoices(
        self,
        year: int,
        month: int,
        status: str | None = None,
    ) -> list:
        """Retrieve persisted invoice records for a month."""
        from apps.invoices.models import InvoiceRecord

        qs = InvoiceRecord.objects.filter(
            tenant=self.tenant,
            billing_date__year=year,
            billing_date__month=month,
        )
        if status:
            qs = qs.filter(status=status)
        return list(qs.order_by("customer_name", "billing_date"))
