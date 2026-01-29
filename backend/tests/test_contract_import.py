"""Tests for contract import service."""

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

from apps.contracts.models import Contract, ContractItem
from apps.contracts.services import (
    ExcelParser,
    ExcelRow,
    ImportProposal,
    ImportService,
    MatchResult,
    MatchStatus,
)
from apps.customers.models import Customer
from apps.products.models import Product


@pytest.fixture
def customer(db, tenant):
    """Create a test customer."""
    return Customer.objects.create(
        tenant=tenant,
        name="EBARA Corporation",
        is_active=True,
    )


@pytest.fixture
def customer_with_netsuite_number(db, tenant):
    """Create a customer with NetSuite customer number."""
    return Customer.objects.create(
        tenant=tenant,
        name="Acme Inc",
        netsuite_customer_number="CUS100",
        is_active=True,
    )


@pytest.fixture
def product(db, tenant):
    """Create a test product."""
    return Product.objects.create(
        tenant=tenant,
        name="Support Service",
        sku="SUPPORT-001",
        type=Product.ProductType.SUBSCRIPTION,
        is_active=True,
    )


@pytest.fixture
def product_with_netsuite_name(db, tenant):
    """Create a product with NetSuite item name."""
    return Product.objects.create(
        tenant=tenant,
        name="Premium Support",
        netsuite_item_name="Annual Support Plan",
        sku="PREMIUM-001",
        type=Product.ProductType.SUBSCRIPTION,
        is_active=True,
    )


def create_test_excel(rows: list[dict]) -> BytesIO:
    """Create a test Excel file with given data rows."""
    wb = Workbook()
    ws = wb.active

    # Headers in row 5
    headers = [
        "Name",
        "Sales Order Number",
        "Contract (2)",
        "Item",
        "Invoicing Instructions",
        "Renewals Exclusion",
        "Contract Start Date",
        "Contract End Date",
        "Sum of List Rate",
        "Sum of Amount",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)

    # Data rows starting at row 6
    for i, row in enumerate(rows, 6):
        ws.cell(row=i, column=1, value=row.get("name"))
        ws.cell(row=i, column=2, value=row.get("sales_order_number"))
        ws.cell(row=i, column=3, value=row.get("contract_number"))
        ws.cell(row=i, column=4, value=row.get("item"))
        ws.cell(row=i, column=5, value=row.get("invoicing_instructions", ""))
        ws.cell(row=i, column=6, value=row.get("renewals_exclusion", ""))
        ws.cell(row=i, column=7, value=row.get("start_date", date(2025, 1, 1)))
        ws.cell(row=i, column=8, value=row.get("end_date", date(2025, 12, 31)))
        ws.cell(row=i, column=9, value=row.get("list_rate", 100.0))
        ws.cell(row=i, column=10, value=row.get("amount", 1200.0))

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class TestExcelRow:
    """Tests for ExcelRow dataclass."""

    def test_parses_customer_number_and_name(self):
        """Test that customer number and name are extracted correctly."""
        row = ExcelRow(
            name="CUS174 EBARA Corporation",
            sales_order_number="SO001",
            contract_number="CON001",
            item="Support",
            invoicing_instructions="",
            renewals_exclusion="",
            contract_start_date=date(2025, 1, 1),
            contract_end_date=date(2025, 12, 31),
            sum_of_list_rate=Decimal("100"),
            sum_of_amount=Decimal("1200"),
        )
        assert row.customer_number == "CUS174"
        assert row.customer_name == "EBARA Corporation"

    def test_handles_name_without_prefix(self):
        """Test handling of name without CUS prefix."""
        row = ExcelRow(
            name="Some Other Name",
            sales_order_number="SO001",
            contract_number="CON001",
            item="Support",
            invoicing_instructions="",
            renewals_exclusion="",
            contract_start_date=date(2025, 1, 1),
            contract_end_date=date(2025, 12, 31),
            sum_of_list_rate=Decimal("100"),
            sum_of_amount=Decimal("1200"),
        )
        assert row.customer_number == ""
        assert row.customer_name == ""


class TestExcelParser:
    """Tests for ExcelParser."""

    def test_parses_valid_excel(self):
        """Test parsing a valid Excel file."""
        excel_file = create_test_excel([
            {
                "name": "CUS001 Test Customer",
                "sales_order_number": "SO001",
                "contract_number": "CON001",
                "item": "Support Service",
                "start_date": date(2025, 1, 1),
                "end_date": date(2025, 12, 31),
                "list_rate": 100.0,
                "amount": 1200.0,
            }
        ])

        parser = ExcelParser()
        rows = parser.parse(excel_file)

        assert len(rows) == 1
        assert rows[0].customer_number == "CUS001"
        assert rows[0].customer_name == "Test Customer"
        assert rows[0].sales_order_number == "SO001"
        assert rows[0].item == "Support Service"
        assert rows[0].sum_of_list_rate == Decimal("100")

    def test_skips_rows_without_cus_prefix(self):
        """Test that rows without CUS prefix are skipped."""
        excel_file = create_test_excel([
            {
                "name": "CUS001 Valid Customer",
                "sales_order_number": "SO001",
                "contract_number": "CON001",
                "item": "Support",
            },
            {
                "name": "Invalid Customer",  # Missing CUS prefix
                "sales_order_number": "SO002",
                "contract_number": "CON002",
                "item": "Support",
            },
        ])

        parser = ExcelParser()
        rows = parser.parse(excel_file)

        assert len(rows) == 1
        assert rows[0].customer_number == "CUS001"

    def test_reports_header_mismatch(self):
        """Test that header mismatches are reported as errors."""
        wb = Workbook()
        ws = wb.active
        # Wrong headers
        ws.cell(row=5, column=1, value="Wrong Header")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        parser = ExcelParser()
        rows = parser.parse(output)

        assert len(rows) == 0
        assert len(parser.errors) > 0
        assert "Header mismatch" in parser.errors[0]


class TestImportService:
    """Tests for ImportService."""

    def test_matches_customer_by_netsuite_number(self, db, tenant, customer_with_netsuite_number):
        """Test that customers are matched by NetSuite number."""
        service = ImportService(tenant)

        result = service._match_customer("Different Name", "CUS100")

        assert result.status == MatchStatus.MATCHED
        assert result.customer == customer_with_netsuite_number
        assert result.confidence == 1.0

    def test_matches_customer_by_name_fuzzy(self, db, tenant, customer):
        """Test fuzzy matching of customer names."""
        service = ImportService(tenant)

        result = service._match_customer("EBARA Corp", "CUS999")

        assert result.status in [MatchStatus.MATCHED, MatchStatus.REVIEW]
        assert result.customer == customer
        assert result.confidence > 0.5

    def test_no_match_returns_not_found(self, db, tenant):
        """Test that unmatched customers return NOT_FOUND."""
        service = ImportService(tenant)

        result = service._match_customer("Completely Different Company", "CUS999")

        assert result.status == MatchStatus.NOT_FOUND
        assert result.customer is None

    def test_matches_product_by_netsuite_name(self, db, tenant, product_with_netsuite_name):
        """Test that products are matched by NetSuite item name."""
        service = ImportService(tenant)

        result = service._match_product("Annual Support Plan")

        assert result == product_with_netsuite_name

    def test_generate_proposals_groups_by_order(self, db, tenant, customer):
        """Test that rows are grouped by sales order number."""
        rows = [
            ExcelRow(
                name="CUS001 EBARA Corporation",
                sales_order_number="SO001",
                contract_number="CON001",
                item="Item 1",
                invoicing_instructions="",
                renewals_exclusion="",
                contract_start_date=date(2025, 1, 1),
                contract_end_date=date(2025, 12, 31),
                sum_of_list_rate=Decimal("100"),
                sum_of_amount=Decimal("1200"),
            ),
            ExcelRow(
                name="CUS001 EBARA Corporation",
                sales_order_number="SO001",
                contract_number="CON001",
                item="Item 2",
                invoicing_instructions="",
                renewals_exclusion="",
                contract_start_date=date(2025, 1, 1),
                contract_end_date=date(2025, 12, 31),
                sum_of_list_rate=Decimal("50"),
                sum_of_amount=Decimal("600"),
            ),
        ]

        service = ImportService(tenant)
        proposals = service.generate_proposals(rows)

        assert len(proposals) == 1
        assert len(proposals[0].items) == 2
        assert proposals[0].total_monthly_rate == Decimal("150")

    def test_discount_rows_are_stored_separately(self, db, tenant, customer):
        """Test that Sales Discount rows are stored in discount_amount."""
        rows = [
            ExcelRow(
                name="CUS001 EBARA Corporation",
                sales_order_number="SO001",
                contract_number="CON001",
                item="Support Service",
                invoicing_instructions="",
                renewals_exclusion="",
                contract_start_date=date(2025, 1, 1),
                contract_end_date=date(2025, 12, 31),
                sum_of_list_rate=Decimal("100"),
                sum_of_amount=Decimal("1200"),
            ),
            ExcelRow(
                name="CUS001 EBARA Corporation",
                sales_order_number="SO001",
                contract_number="CON001",
                item="Sales Discount",
                invoicing_instructions="",
                renewals_exclusion="",
                contract_start_date=date(2025, 1, 1),
                contract_end_date=date(2025, 12, 31),
                sum_of_list_rate=Decimal("-10"),
                sum_of_amount=Decimal("-120"),
            ),
        ]

        service = ImportService(tenant)
        proposals = service.generate_proposals(rows)

        assert len(proposals) == 1
        assert len(proposals[0].items) == 1  # Discount not in items
        assert proposals[0].discount_amount == Decimal("-10")
        assert proposals[0].total_monthly_rate == Decimal("90")  # 100 - 10

    def test_apply_proposals_creates_contracts(self, db, tenant, customer):
        """Test that approved proposals create contracts."""
        service = ImportService(tenant)

        proposal = ImportProposal(
            customer_number="CUS001",
            customer_name="EBARA Corporation",
            sales_order_number="SO001",
            contract_number="CON001",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            approved=True,
            selected_customer=customer,
        )
        proposal.items.append(
            service._match_product("Support") or
            type("ContractLineItem", (), {
                "item_name": "Support Service",
                "monthly_rate": Decimal("100"),
                "product": None,
            })()
        )

        # Actually create the ContractLineItem properly
        from apps.contracts.services.import_service import ContractLineItem
        proposal.items = [ContractLineItem(
            item_name="Support Service",
            monthly_rate=Decimal("100"),
            product=None,
        )]

        service.proposals = [proposal]
        contracts = service.apply_proposals(auto_create_products=True)

        assert len(contracts) == 1
        assert contracts[0].customer == customer
        assert contracts[0].netsuite_sales_order_number == "SO001"

        # Check that product was auto-created
        product = Product.objects.filter(netsuite_item_name="Support Service").first()
        assert product is not None

        # Check contract item
        items = ContractItem.objects.filter(contract=contracts[0])
        assert items.count() == 1
        assert items[0].product == product

    def test_apply_proposals_skips_rejected(self, db, tenant, customer):
        """Test that rejected proposals are not created."""
        service = ImportService(tenant)

        proposal = ImportProposal(
            customer_number="CUS001",
            customer_name="EBARA Corporation",
            sales_order_number="SO001",
            contract_number="CON001",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            approved=False,
            rejected=True,
            selected_customer=customer,
        )

        service.proposals = [proposal]
        contracts = service.apply_proposals()

        assert len(contracts) == 0

    def test_apply_proposals_requires_customer(self, db, tenant):
        """Test that proposals without selected customer fail."""
        service = ImportService(tenant)

        proposal = ImportProposal(
            customer_number="CUS001",
            customer_name="Unknown Customer",
            sales_order_number="SO001",
            contract_number="CON001",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            approved=True,
            selected_customer=None,
        )

        service.proposals = [proposal]
        contracts = service.apply_proposals()

        assert len(contracts) == 0
        assert proposal.error == "No customer selected"

    def test_get_summary(self, db, tenant, customer):
        """Test summary calculation."""
        service = ImportService(tenant)

        # Create proposals with different statuses
        service.proposals = [
            ImportProposal(
                customer_number="CUS001",
                customer_name="Customer 1",
                sales_order_number="SO001",
                match_result=MatchResult(
                    status=MatchStatus.MATCHED,
                    customer=customer,
                    confidence=0.95,
                ),
            ),
            ImportProposal(
                customer_number="CUS002",
                customer_name="Customer 2",
                sales_order_number="SO002",
                match_result=MatchResult(
                    status=MatchStatus.REVIEW,
                    confidence=0.75,
                ),
            ),
            ImportProposal(
                customer_number="CUS003",
                customer_name="Customer 3",
                sales_order_number="SO003",
                match_result=MatchResult(
                    status=MatchStatus.NOT_FOUND,
                ),
            ),
        ]

        summary = service.get_summary()

        assert summary["total_proposals"] == 3
        assert summary["auto_matched"] == 1
        assert summary["needs_review"] == 1
        assert summary["not_found"] == 1


class TestImportProposal:
    """Tests for ImportProposal dataclass."""

    def test_total_monthly_rate_calculation(self):
        """Test total monthly rate includes items and discount."""
        from apps.contracts.services.import_service import ContractLineItem

        proposal = ImportProposal(
            customer_number="CUS001",
            customer_name="Test",
            discount_amount=Decimal("-20"),
        )
        proposal.items = [
            ContractLineItem(item_name="Item 1", monthly_rate=Decimal("100")),
            ContractLineItem(item_name="Item 2", monthly_rate=Decimal("50")),
        ]

        assert proposal.total_monthly_rate == Decimal("130")  # 100 + 50 - 20

    def test_needs_review_flag(self):
        """Test needs_review property."""
        proposal_review = ImportProposal(
            customer_number="CUS001",
            customer_name="Test",
            match_result=MatchResult(status=MatchStatus.REVIEW),
        )
        assert proposal_review.needs_review is True

        proposal_matched = ImportProposal(
            customer_number="CUS001",
            customer_name="Test",
            match_result=MatchResult(status=MatchStatus.MATCHED),
        )
        assert proposal_matched.needs_review is False
