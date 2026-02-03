"""Tests for invoice generation service and GraphQL API."""
import pytest
from datetime import date
from decimal import Decimal

from apps.contracts.models import Contract, ContractItem
from apps.customers.models import Customer
from apps.invoices.services import InvoiceService
from apps.invoices.schema import InvoiceQuery, _convert_invoice
from apps.products.models import Product


@pytest.fixture
def customer(db, tenant):
    """Create a test customer."""
    return Customer.objects.create(
        tenant=tenant,
        name="Test Customer",
        address={"street": "123 Main St", "city": "Berlin"},
        is_active=True,
    )


@pytest.fixture
def product(db, tenant):
    """Create a test product."""
    return Product.objects.create(
        tenant=tenant,
        name="Test Product",
        sku="TEST-001",
    )


@pytest.fixture
def monthly_contract(db, tenant, customer):
    """Create an active monthly contract."""
    return Contract.objects.create(
        tenant=tenant,
        customer=customer,
        name="Monthly Contract",
        status=Contract.Status.ACTIVE,
        start_date=date(2026, 1, 1),
        billing_start_date=date(2026, 1, 1),
        billing_interval=Contract.BillingInterval.MONTHLY,
        billing_anchor_day=1,
    )


@pytest.fixture
def quarterly_contract(db, tenant, customer):
    """Create an active quarterly contract."""
    return Contract.objects.create(
        tenant=tenant,
        customer=customer,
        name="Quarterly Contract",
        status=Contract.Status.ACTIVE,
        start_date=date(2026, 1, 1),
        billing_start_date=date(2026, 1, 1),
        billing_interval=Contract.BillingInterval.QUARTERLY,
        billing_anchor_day=1,
    )


class TestInvoiceServiceGetInvoicesForMonth:
    """Test InvoiceService.get_invoices_for_month()."""

    def test_returns_invoices_for_active_contracts(
        self, tenant, monthly_contract, product
    ):
        """Test that invoices are returned for active contracts."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=2,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        assert len(invoices) == 1
        invoice = invoices[0]
        assert invoice.contract_name == "Monthly Contract"
        assert invoice.customer_name == "Test Customer"
        assert invoice.billing_date == date(2026, 1, 1)
        assert len(invoice.line_items) == 1
        assert invoice.total_amount == Decimal("200.00")  # 2 x 100

    def test_excludes_draft_contracts(self, tenant, customer, product):
        """Test that draft contracts are excluded."""
        draft_contract = Contract.objects.create(
            tenant=tenant,
            customer=customer,
            name="Draft Contract",
            status=Contract.Status.DRAFT,
            start_date=date(2026, 1, 1),
            billing_start_date=date(2026, 1, 1),
            billing_interval=Contract.BillingInterval.MONTHLY,
        )
        ContractItem.objects.create(
            tenant=tenant,
            contract=draft_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        assert len(invoices) == 0

    def test_excludes_paused_contracts(self, tenant, customer, product):
        """Test that paused contracts are excluded."""
        paused_contract = Contract.objects.create(
            tenant=tenant,
            customer=customer,
            name="Paused Contract",
            status=Contract.Status.PAUSED,
            start_date=date(2026, 1, 1),
            billing_start_date=date(2026, 1, 1),
            billing_interval=Contract.BillingInterval.MONTHLY,
        )
        ContractItem.objects.create(
            tenant=tenant,
            contract=paused_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        assert len(invoices) == 0

    def test_empty_month_returns_empty_list(self, tenant, monthly_contract, product):
        """Test that a month with no billing events returns empty list."""
        # Contract starts in Jan 2026, so Dec 2025 should be empty
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2025, 12)

        assert len(invoices) == 0

    def test_quarterly_billing_only_in_quarter_months(
        self, tenant, quarterly_contract, product
    ):
        """Test quarterly contracts only generate invoices in quarter months."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=quarterly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("300.00"),
        )

        service = InvoiceService(tenant)

        # January should have invoice (quarterly billing)
        jan_invoices = service.get_invoices_for_month(2026, 1)
        assert len(jan_invoices) == 1

        # February should have no invoice
        feb_invoices = service.get_invoices_for_month(2026, 2)
        assert len(feb_invoices) == 0

        # April should have invoice (next quarter)
        apr_invoices = service.get_invoices_for_month(2026, 4)
        assert len(apr_invoices) == 1

    def test_invoice_includes_customer_address(
        self, tenant, monthly_contract, product
    ):
        """Test that invoice includes customer address."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        assert invoices[0].customer_address == {"street": "123 Main St", "city": "Berlin"}

    def test_invoices_sorted_by_customer_name(self, tenant, product):
        """Test that invoices are sorted by customer name."""
        customer_z = Customer.objects.create(
            tenant=tenant, name="Zebra Corp", is_active=True
        )
        customer_a = Customer.objects.create(
            tenant=tenant, name="Alpha Inc", is_active=True
        )

        contract_z = Contract.objects.create(
            tenant=tenant,
            customer=customer_z,
            name="Contract Z",
            status=Contract.Status.ACTIVE,
            start_date=date(2026, 1, 1),
            billing_start_date=date(2026, 1, 1),
            billing_interval=Contract.BillingInterval.MONTHLY,
        )
        contract_a = Contract.objects.create(
            tenant=tenant,
            customer=customer_a,
            name="Contract A",
            status=Contract.Status.ACTIVE,
            start_date=date(2026, 1, 1),
            billing_start_date=date(2026, 1, 1),
            billing_interval=Contract.BillingInterval.MONTHLY,
        )

        for contract in [contract_z, contract_a]:
            ContractItem.objects.create(
                tenant=tenant,
                contract=contract,
                product=product,
                quantity=1,
                unit_price=Decimal("100.00"),
            )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        assert len(invoices) == 2
        assert invoices[0].customer_name == "Alpha Inc"
        assert invoices[1].customer_name == "Zebra Corp"

    def test_one_off_item_only_billed_once(self, tenant, monthly_contract, product):
        """Test that one-off items only appear in one invoice."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("500.00"),
            is_one_off=True,
            billing_start_date=date(2026, 1, 15),
        )

        service = InvoiceService(tenant)

        # January should have the one-off
        jan_invoices = service.get_invoices_for_month(2026, 1)
        assert len(jan_invoices) == 1
        assert jan_invoices[0].line_items[0].is_one_off is True

        # February should not have the one-off
        feb_invoices = service.get_invoices_for_month(2026, 2)
        # The contract might still have regular items, but no one-off
        for invoice in feb_invoices:
            for item in invoice.line_items:
                assert item.is_one_off is False

    def test_tenant_isolation(self, db, tenant, monthly_contract, product):
        """Test that invoices are scoped to tenant."""
        # Create item for our tenant's contract
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        # Create another tenant with a contract
        from apps.tenants.models import Tenant

        other_tenant = Tenant.objects.create(name="Other Company", currency="USD")
        other_customer = Customer.objects.create(
            tenant=other_tenant, name="Other Customer", is_active=True
        )
        other_product = Product.objects.create(
            tenant=other_tenant, name="Other Product", sku="OTHER-001"
        )
        other_contract = Contract.objects.create(
            tenant=other_tenant,
            customer=other_customer,
            name="Other Contract",
            status=Contract.Status.ACTIVE,
            start_date=date(2026, 1, 1),
            billing_start_date=date(2026, 1, 1),
            billing_interval=Contract.BillingInterval.MONTHLY,
        )
        ContractItem.objects.create(
            tenant=other_tenant,
            contract=other_contract,
            product=other_product,
            quantity=1,
            unit_price=Decimal("999.00"),
        )

        # Our tenant should only see its own invoice
        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        assert len(invoices) == 1
        assert invoices[0].customer_name == "Test Customer"
        assert invoices[0].total_amount == Decimal("100.00")


class TestInvoiceGraphQLConversion:
    """Test GraphQL type conversions."""

    def test_convert_invoice_to_graphql_type(
        self, tenant, monthly_contract, product
    ):
        """Test conversion from InvoiceData to InvoiceType."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=2,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        graphql_invoice = _convert_invoice(invoices[0])

        assert graphql_invoice.contract_id == monthly_contract.id
        assert graphql_invoice.contract_name == "Monthly Contract"
        assert graphql_invoice.customer_name == "Test Customer"
        assert graphql_invoice.total_amount == Decimal("200.00")
        assert graphql_invoice.line_item_count == 1
        assert len(graphql_invoice.line_items) == 1
        assert graphql_invoice.line_items[0].product_name == "Test Product"
        assert graphql_invoice.line_items[0].quantity == 2
        assert graphql_invoice.line_items[0].amount == Decimal("200.00")


class TestInvoicePDFGeneration:
    """Test PDF generation methods."""

    def test_generate_pdf_returns_bytes(self, tenant, monthly_contract, product):
        """Test that generate_pdf returns PDF bytes."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        pdf_bytes = service.generate_pdf(invoices, language="de")

        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 0
        # PDF files start with %PDF
        assert pdf_bytes[:4] == b"%PDF"

    def test_generate_pdf_empty_list_returns_empty(self, tenant):
        """Test that empty invoice list returns empty bytes."""
        service = InvoiceService(tenant)
        pdf_bytes = service.generate_pdf([], language="de")
        assert pdf_bytes == b""

    def test_generate_individual_pdfs_returns_zip(
        self, tenant, monthly_contract, product
    ):
        """Test that generate_individual_pdfs returns a ZIP file."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        zip_bytes = service.generate_individual_pdfs(invoices, 2026, 1, language="en")

        assert isinstance(zip_bytes, bytes)
        assert len(zip_bytes) > 0
        # ZIP files start with PK
        assert zip_bytes[:2] == b"PK"

    def test_generate_individual_pdfs_contains_correct_files(
        self, tenant, monthly_contract, product
    ):
        """Test that ZIP contains correctly named PDF files."""
        import zipfile
        import io

        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        zip_bytes = service.generate_individual_pdfs(invoices, 2026, 1, language="de")

        with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
            filenames = zf.namelist()
            assert len(filenames) == 1
            assert "invoice-Test-Customer-Monthly-Contract-2026-01.pdf" in filenames

            # Verify content is a valid PDF
            pdf_content = zf.read(filenames[0])
            assert pdf_content[:4] == b"%PDF"

    def test_pdf_language_support(self, tenant, monthly_contract, product):
        """Test that PDF generation works for both languages."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)

        # Both languages should produce valid PDFs
        pdf_de = service.generate_pdf(invoices, language="de")
        pdf_en = service.generate_pdf(invoices, language="en")

        assert pdf_de[:4] == b"%PDF"
        assert pdf_en[:4] == b"%PDF"
        # German and English PDFs should be different (different labels)
        assert pdf_de != pdf_en


class TestInvoiceExcelGeneration:
    """Test Excel generation methods."""

    def test_generate_excel_returns_bytes(self, tenant, monthly_contract, product):
        """Test that generate_excel returns Excel bytes."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        excel_bytes = service.generate_excel(invoices, 2026, 1, language="de")

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0
        # Excel files (xlsx) start with PK (ZIP format)
        assert excel_bytes[:2] == b"PK"

    def test_generate_excel_has_three_sheets(self, tenant, monthly_contract, product):
        """Test that Excel file contains three sheets."""
        import io
        from openpyxl import load_workbook

        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        excel_bytes = service.generate_excel(invoices, 2026, 1, language="en")

        wb = load_workbook(io.BytesIO(excel_bytes))
        assert len(wb.sheetnames) == 2
        assert "Summary" in wb.sheetnames
        assert "Details" in wb.sheetnames

    def test_excel_summary_sheet_contains_totals(
        self, tenant, monthly_contract, product
    ):
        """Test that Summary sheet contains correct totals."""
        import io
        from openpyxl import load_workbook

        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=2,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        excel_bytes = service.generate_excel(invoices, 2026, 1, language="de")

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Summary"]

        # Row 3 has headers, row 4 has first data row, row 5 has totals
        # Column J (10) has the Amount
        assert ws["A4"].value == "Test Customer"  # Customer name
        assert ws["J4"].value == 200.0  # Amount (2 x 100)
        # Total row
        assert ws["A5"].value == "Gesamtbetrag"  # Total label in German
        assert ws["J5"].value == 200.0  # Total amount

    def test_excel_summary_sheet_contains_contract_data(
        self, tenant, monthly_contract, product
    ):
        """Test that Summary sheet contains contract data."""
        import io
        from openpyxl import load_workbook

        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        excel_bytes = service.generate_excel(invoices, 2026, 1, language="en")

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Summary"]

        # Row 3 has headers, row 4 has first data row
        # Column A: Customer, Column J: Amount
        assert ws["A4"].value == "Test Customer"
        assert ws["J4"].value == 100.0

    def test_excel_details_sheet_contains_data(
        self, tenant, monthly_contract, product
    ):
        """Test that Details sheet contains line item data."""
        import io
        from openpyxl import load_workbook

        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=3,
            unit_price=Decimal("50.00"),
        )

        service = InvoiceService(tenant)
        invoices = service.get_invoices_for_month(2026, 1)
        excel_bytes = service.generate_excel(invoices, 2026, 1, language="en")

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Details"]

        # Row 1 is header, row 2 has first line item
        # Column A: Customer, F: Item, R: Quantity, S: Unit Price, T: Amount
        assert ws["A2"].value == "Test Customer"
        assert ws["F2"].value == "Test Product"
        assert ws["R2"].value == 3  # quantity
        assert ws["S2"].value == 50.0  # unit price
        assert ws["T2"].value == 150.0  # amount (3 x 50)

    def test_generate_excel_empty_list(self, tenant):
        """Test that empty invoice list returns valid Excel file."""
        import io
        from openpyxl import load_workbook

        service = InvoiceService(tenant)
        excel_bytes = service.generate_excel([], 2026, 1, language="de")

        assert len(excel_bytes) > 0
        # Should still be a valid Excel file
        wb = load_workbook(io.BytesIO(excel_bytes))
        assert len(wb.sheetnames) >= 1


class TestInvoiceExportEndpoint:
    """Test the REST export endpoint."""

    @pytest.fixture
    def auth_headers(self, user):
        """Create JWT auth headers for the test user."""
        from apps.core.auth import create_access_token

        token = create_access_token(user)
        return {"HTTP_AUTHORIZATION": f"Bearer {token}"}

    def test_export_requires_authentication(self, client):
        """Test that export endpoint requires auth."""
        response = client.get("/api/invoices/export/?year=2026&month=1&format=pdf")
        assert response.status_code == 401

    def test_export_requires_year_and_month(self, client, auth_headers):
        """Test that year and month are required."""
        # Missing year
        response = client.get("/api/invoices/export/?month=1&format=pdf", **auth_headers)
        assert response.status_code == 400

        # Missing month
        response = client.get("/api/invoices/export/?year=2026&format=pdf", **auth_headers)
        assert response.status_code == 400

    def test_export_validates_month_range(self, client, auth_headers):
        """Test that month must be 1-12."""
        response = client.get("/api/invoices/export/?year=2026&month=13&format=pdf", **auth_headers)
        assert response.status_code == 400

        response = client.get("/api/invoices/export/?year=2026&month=0&format=pdf", **auth_headers)
        assert response.status_code == 400

    def test_export_validates_format(self, client, auth_headers):
        """Test that format must be valid."""
        response = client.get("/api/invoices/export/?year=2026&month=1&format=invalid", **auth_headers)
        assert response.status_code == 400

    def test_export_returns_404_when_no_invoices(self, client, auth_headers):
        """Test that endpoint returns 404 when no invoices exist."""
        response = client.get("/api/invoices/export/?year=2020&month=1&format=pdf", **auth_headers)
        assert response.status_code == 404

    def test_export_pdf_success(
        self, client, auth_headers, tenant, monthly_contract, product
    ):
        """Test successful PDF export."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        response = client.get("/api/invoices/export/?year=2026&month=1&format=pdf", **auth_headers)

        assert response.status_code == 200
        assert response["Content-Type"] == "application/pdf"
        assert 'filename="invoices-2026-01.pdf"' in response["Content-Disposition"]
        assert response.content[:4] == b"%PDF"

    def test_export_excel_success(
        self, client, auth_headers, tenant, monthly_contract, product
    ):
        """Test successful Excel export."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        response = client.get("/api/invoices/export/?year=2026&month=1&format=excel", **auth_headers)

        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]
        assert 'filename="invoices-2026-01.xlsx"' in response["Content-Disposition"]
        # Excel (xlsx) files start with PK (ZIP format)
        assert response.content[:2] == b"PK"

    def test_export_individual_pdfs_success(
        self, client, auth_headers, tenant, monthly_contract, product
    ):
        """Test successful individual PDFs export as ZIP."""
        ContractItem.objects.create(
            tenant=tenant,
            contract=monthly_contract,
            product=product,
            quantity=1,
            unit_price=Decimal("100.00"),
        )

        response = client.get(
            "/api/invoices/export/?year=2026&month=1&format=pdf-individual", **auth_headers
        )

        assert response.status_code == 200
        assert response["Content-Type"] == "application/zip"
        assert 'filename="invoices-2026-01.zip"' in response["Content-Disposition"]
        # ZIP files start with PK
        assert response.content[:2] == b"PK"
